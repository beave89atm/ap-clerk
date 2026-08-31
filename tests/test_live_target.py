"""Live target is explicit and write-blocked. No network I/O; no secrets printed."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from ap_clerk.auth import load_credentials, resolve_target
from ap_clerk.cli import main
from ap_clerk.kimco import (
    LIVE_SERVICES,
    PROTOTYPE_SERVICES,
    KimcoClient,
    KimcoError,
    services_for,
)

LIVE_URL = "https://live.kimcoerp.com"
PROTO_URL = "https://prototype.kimcoerp.com"
FIXTURE = Path(__file__).resolve().parent.parent / "fixtures" / "testrun-727-803.json"


def _clear_kimco_env(monkeypatch: pytest.MonkeyPatch) -> None:
    for name in (
        "KIMCO_PROTOTYPE_API_KEY",
        "KIMCO_PROTOTYPE_API_PASSWORD",
        "KIMCO_PROTOTYPE_INSTANCE_URL",
        "KIMCO_API_KEY",
        "KIMCO_API_PASSWORD",
        "KIMCO_LIVE_API_KEY",
        "KIMCO_LIVE_API_PASSWORD",
        "KIMCO_LIVE_INSTANCE_URL",
        "KIMCO_TARGET",
    ):
        monkeypatch.delenv(name, raising=False)


def test_default_target_is_prototype(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    assert resolve_target(live_flag=False) == "prototype"
    assert services_for("prototype") == PROTOTYPE_SERVICES


def test_live_flag_and_env_select_live(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    assert resolve_target(live_flag=True) == "live"
    monkeypatch.setenv("KIMCO_TARGET", "live")
    assert resolve_target(live_flag=False) == "live"
    assert services_for("live") == LIVE_SERVICES
    assert LIVE_SERVICES["ap_invoices"] == "bcca4094b6ec4564942b19f5d7bb255c"
    assert LIVE_SERVICES["ap_batches"] == "31bf524dcd5b464580d4a1b55c01881e"
    assert LIVE_SERVICES["purchase_lines"] == "f1f8732f8daa4e2b9d8065037f7bb43d"
    assert LIVE_SERVICES["receipts"] == "494eafafa31a42bba7eb8697a36a3f0a"


def test_prototype_ignores_live_creds(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    monkeypatch.setenv("KIMCO_LIVE_API_KEY", "live-key")
    monkeypatch.setenv("KIMCO_LIVE_API_PASSWORD", "live-pw")
    creds = load_credentials(target="prototype")
    assert creds.target == "prototype"
    assert creds.ready is False
    assert creds.key is None
    assert creds.instance_url == PROTO_URL


def test_live_refuses_without_live_creds(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    monkeypatch.setenv("KIMCO_PROTOTYPE_API_KEY", "proto-key")
    monkeypatch.setenv("KIMCO_PROTOTYPE_API_PASSWORD", "proto-pw")
    creds = load_credentials(target="live")
    assert creds.ready is False
    assert creds.key is None
    assert creds.password is None
    assert "KIMCO_LIVE_API_KEY" in (creds.error or "")
    assert creds.instance_url == LIVE_URL


def test_live_uses_only_live_creds(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    monkeypatch.setenv("KIMCO_LIVE_API_KEY", "live-key")
    monkeypatch.setenv("KIMCO_LIVE_API_PASSWORD", "live-pw")
    monkeypatch.setenv("KIMCO_PROTOTYPE_API_KEY", "proto-key")
    monkeypatch.setenv("KIMCO_PROTOTYPE_API_PASSWORD", "proto-pw")
    creds = load_credentials(target="live")
    assert creds.ready is True
    assert creds.key == "live-key"
    assert creds.key_source == "KIMCO_LIVE_API_KEY/KIMCO_LIVE_API_PASSWORD"
    assert creds.instance_url == LIVE_URL


def test_live_refuses_non_live_instance_url(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    monkeypatch.setenv("KIMCO_LIVE_API_KEY", "live-key")
    monkeypatch.setenv("KIMCO_LIVE_API_PASSWORD", "live-pw")
    monkeypatch.setenv("KIMCO_LIVE_INSTANCE_URL", PROTO_URL)
    creds = load_credentials(target="live")
    assert creds.ready is False
    assert creds.key is None
    assert "live.kimcoerp.com" in (creds.error or "")


def test_prototype_still_refuses_live_url(monkeypatch: pytest.MonkeyPatch) -> None:
    _clear_kimco_env(monkeypatch)
    monkeypatch.setenv("KIMCO_PROTOTYPE_API_KEY", "proto-key")
    monkeypatch.setenv("KIMCO_PROTOTYPE_API_PASSWORD", "proto-pw")
    monkeypatch.setenv("KIMCO_PROTOTYPE_INSTANCE_URL", LIVE_URL)
    creds = load_credentials(target="prototype")
    assert creds.ready is False
    assert creds.key is None


def test_prototype_client_refuses_live_host() -> None:
    with pytest.raises(KimcoError, match="live.kimcoerp.com"):
        KimcoClient(LIVE_URL, "token")


def test_live_client_allows_writes_on_live_host() -> None:
    client = KimcoClient(LIVE_URL, "token", target="live")
    assert client.services["ap_invoices"] == LIVE_SERVICES["ap_invoices"]

    class FakeResp:
        status_code = 200
        text = '{"id": 1}'
        headers = {}

        def json(self):
            return {"id": 1}

    with patch.object(client.session, "request", return_value=FakeResp()) as req:
        created_id, _body, status, error = client.create("ap_invoices", {"Invoice_Number": "x"})
        assert created_id == 1
        assert status == 200
        assert error == ""
        req.assert_called()
        assert req.call_args.args[0] == "POST"


def test_live_client_refuses_prototype_host() -> None:
    with pytest.raises(KimcoError, match="live.kimcoerp.com"):
        KimcoClient(PROTO_URL, "token", target="live")
    client = KimcoClient(LIVE_URL, "token", target="live")
    with pytest.raises(KimcoError, match="non-live"):
        client.request("GET", f"{PROTO_URL}/api/v2/{LIVE_SERVICES['ap_invoices']}")


def test_cli_live_refuses_without_creds(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    _clear_kimco_env(monkeypatch)
    report = tmp_path / "hold.xlsx"
    code = main(
        [
            "enter",
            "--live",
            "--fixture",
            str(FIXTURE),
            "--report",
            str(report),
            "--as-of",
            "2026-08-28",
        ]
    )
    assert code == 2
    out = capsys.readouterr().out
    assert "KIMCO_LIVE_API_KEY: absent" in out
    assert "Target: live" in out
    assert "token" not in out.lower() or "token not printed" in out.lower()
    assert report.exists()


def test_cli_live_auth_success_runs_enter(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    _clear_kimco_env(monkeypatch)
    monkeypatch.setenv("KIMCO_LIVE_API_KEY", "live-key")
    monkeypatch.setenv("KIMCO_LIVE_API_PASSWORD", "live-pw")
    report = tmp_path / "live.xlsx"

    class FakeClient:
        target = "live"

    def fake_auth(base_url: str, key: str, password: str, *, target: str = "prototype") -> FakeClient:
        assert target == "live"
        assert base_url == LIVE_URL
        assert key == "live-key"
        return FakeClient()

    def fake_enter(client, invoices, **kwargs):
        assert client.target == "live"
        assert kwargs.get("flag_outlook") is True
        return [
            {
                "Vendor": "Test Vendor",
                "Invoice #": "1",
                "date": "2026-08-28",
                "PO": "",
                "Amount": 1,
                "Result": "Success",
                "Why": "Header created",
                "KIMCO id": 100,
                "Batch": "API Agent - 8/28/26 (1)",
                "Fees and surcharges": "none",
                "PPV": "none",
                "Attach status": "no-pdf-on-vm",
                "Flag in Outlook": "Yes",
            }
        ]

    with patch("ap_clerk.cli.KimcoClient.authenticate", side_effect=fake_auth):
        with patch("ap_clerk.cli.run_enter", side_effect=fake_enter) as run_enter:
            code = main(
                [
                    "enter",
                    "--live",
                    "--fixture",
                    str(FIXTURE),
                    "--report",
                    str(report),
                    "--as-of",
                    "2026-08-28",
                ]
            )
            run_enter.assert_called_once()
    assert code == 0
    out = capsys.readouterr().out
    assert "Live auth success" in out
    assert "token not printed" in out
    assert "Entered in AI" in out
    assert "AI HOLD" in out
    assert "follow-up flag" in out.lower() or "No follow-up flag" in out
    assert report.exists()
    sheet = load_workbook(report).active
    assert sheet.cell(1, 13).value == "Flag in Outlook"
    assert sheet.cell(2, 6).value == "Success"
    assert sheet.cell(2, 13).value == "Yes"
