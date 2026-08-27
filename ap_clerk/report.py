"""Excel run report writer."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Vendor",
    "Invoice #",
    "date",
    "PO",
    "Amount",
    "Result",
    "Why",
    "KIMCO id",
    "Batch",
    "Fees and surcharges",
    "PPV",
    "Attach status",
]


def write_report(path: Path, rows: list[dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP run"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(1, col, name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)
    fills = {
        "Success": PatternFill("solid", fgColor="C6EFCE"),
        "Fail": PatternFill("solid", fgColor="FFC7CE"),
        "HOLD": PatternFill("solid", fgColor="FFEB9C"),
    }
    for row_idx, row in enumerate(rows, start=2):
        values = [row.get(col, "") for col in COLUMNS]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row_idx, col, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if COLUMNS[col - 1] == "Result":
                fill = fills.get(str(value))
                if fill:
                    cell.fill = fill
    widths = [28, 18, 12, 12, 12, 10, 55, 12, 22, 40, 10, 16]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, len(rows) + 1)}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22
    workbook.save(path)
    return path
